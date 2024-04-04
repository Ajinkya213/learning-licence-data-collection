from MCQ import MCQ
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils import get_column_letter
import json
import os

# Load the workbook
curr_dir = os.getcwd()
xls_file_path = os.path.join(curr_dir,"question-bank.xlsx")
workbook = load_workbook(xls_file_path)
sheet = workbook.active

# List of MCQ objects
mcq_list = []

# Directory to save images
resource_dir = "resource"
os.makedirs(resource_dir, exist_ok=True)

#Loading images from the spreadsheet
image_loader = SheetImageLoader(sheet)

# Iterate through rows in the worksheet
for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7), start=2):
    srno = row[0].value
    # Remove newline characters, tabs, and whitespace from question
    question = str(row[1].value).replace('\n', '').replace('\t', '').strip()
    # Remove newline characters, tabs, and whitespace from options
    options = [str(row[i].value).capitalize().replace('\n', '').replace('\t', '').strip() if isinstance(row[i].value, str) else row[i].value for i in range(2, 5)]
    answer = row[5].value
    
    # Construct the cell reference for the image
    image_cell_ref = f'{get_column_letter(7)}{row_idx}'  # 7 corresponds to column G
    
    if image_loader.image_in(image_cell_ref):
        # Get the image from the cell
        image = image_loader.get(image_cell_ref)
        
        # Convert the image to RGB mode
        image = image.convert('RGB')
        
        # Save the image with the serial number as the filename in the resource directory
        image_filename = f"{srno}.jpg"
        image_path = os.path.join(resource_dir, image_filename)
        image.save(image_path)
        
        # Update the image attribute to the location of the saved image
        image_location = os.path.relpath(image_path, os.getcwd())  # Get the relative path
        image_location= '/'+image_location
    else:
        image_location = None
    
    # Create an instance of the MCQ class
    mcq = MCQ(srno=srno, question=question, options=options, answer=answer, image=image_location)
    
    # Append the MCQ object to the list
    mcq_list.append(mcq)

# Close the workbook
workbook.close()

# Now you have a list of MCQ objects representing the data from your Excel sheet
for mcq in mcq_list:
    print("Question:", mcq.question)
    print("Options:", mcq.options)
    print("Answer:", mcq.answer)
    print("Image Location:", mcq.image)  # Now contains the location of the saved image in the resource directory
    print()

json_file_path = "mcq_data.json"

# Convert list of MCQ objects to dictionary
mcq_list_dict = [mcq.__dict__ for mcq in mcq_list]

# Write the data to a JSON file
with open(json_file_path, 'w') as json_file:
    json.dump(mcq_list_dict, json_file, indent=4)

print("MCQ data saved to:", json_file_path)